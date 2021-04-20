#Excel automation with openpyxl
from openpyxl import load_workbook

wb = load_workbook('Financial Sample.xlsx')
ws1 = wb['Finances 2017']
maxrow = ws1.max_row


def insert_sum():
    ws1['L' + str(maxrow)] = "=SUM(L2:L" + str(maxrow - 1) + ")"


if __name__ == "__main__":
    insert_sum()
    wb.save('Financial Sample.xlsx')

#openpyxl commands
wb = load_workbook('Financial Sample.xlsx')

wb.sheetnames

ws1 = wb.active

ws1['C9'].value

ws1['B9'].value

for row in range(2, ws1.max_row):
    cell = 'B' + str(row)
    print(ws1[cell].value)

#####

profit_total = 0
for col in list('L'):
    for row in range(2, 101):
        cell = col + str(row)
        profit_total += float(ws1[cell].value)

print(profit_total)

#####


wb = load_workbook('Financial Sample.xlsx')
ws1 = wb['Finances 2017']

ws1['L703'] = "=SUM(L2:L701)"
wb.save('Financial Sample.xlsx')

ws1['L' + str(ws1.max_row)] = "=SUM(L2:L" + str(ws1.max_row - 1) + ")"
wb.save('Financial Sample.xlsx')